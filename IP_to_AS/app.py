import tkinter as tk
from tkinter import filedialog
import openpyxl
import requests
import os
import webbrowser

def get_ip_info(ip):
    url = f"http://ip-api.com/json/{ip}"
    response = requests.get(url)
    data = response.json()
    return data.get('as'), data.get('country')

def process_excel_file(file_path):
    if file_path:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in range(2, sheet.max_row + 1):  # Assuming data starts from row 2
            ip_address = sheet.cell(row=row, column=1).value
            if ip_address:
                as_info, country = get_ip_info(ip_address)
                sheet.cell(row=row, column=2).value = as_info
                sheet.cell(row=row, column=3).value = country

        workbook.save(file_path)
        workbook.close()
        os.startfile(file_path)
        print("Processing complete.")

def select_file():
    file_path = filedialog.askopenfilename()
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def on_process():
    file_path = entry.get()
    if file_path:
        process_excel_file(file_path)

def open_link(event):
    webbrowser.open_new("https://github.com/luudinhmac/Tools-Network")

def main():
    root = tk.Tk()
    root.title("IP Info Processor")
    root.geometry("450x200")  # Set size to 15cm x 5cm

    button_select = tk.Button(root, text="Select Excel File", command=select_file)
    button_select.pack(pady=10)

    global entry
    entry = tk.Entry(root, width=50)
    entry.pack(pady=5)

    button_process = tk.Button(root, text="THỰC HIỆN", command=on_process)
    button_process.pack(pady=10)

    link = tk.Label(root, text="More tools", fg="blue", cursor="hand2")
    link.pack(pady=10)
    link.bind("<Button-1>", open_link)

    root.mainloop()

if __name__ == "__main__":
    main()
